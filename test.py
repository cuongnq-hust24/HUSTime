import openpyxl
import datetime

tuan1 = datetime.date(2024, 9, 2)

print(tuan1 + datetime.timedelta(weeks=1))
print(tuan1.strftime("%m/%d/%Y"))

workbook = openpyxl.load_workbook('tkb.xlsx')
sheets = workbook.sheetnames
tkb = workbook[sheets[0]]
danhsachmon = workbook[sheets[1]]

print(tkb.cell(row=9, column=5).value.split(','))
