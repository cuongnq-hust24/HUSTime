import pandas as pd
import openpyxl
import datetime

data = {'Subject': [], 'Start Date': [], 'End Date': [], 'Start Time': [], 'End Time': [], 'Description': [],
        'Location': []}
# data = pd.DataFrame(form)

workbook = openpyxl.load_workbook('tkb.xlsx')
sheets = workbook.sheetnames
tkb = workbook[sheets[0]]
danhsachmon = workbook[sheets[1]]

tuan1 = datetime.date(2024, 9, 2)


def get_name(class_id):
    found = False
    for row in danhsachmon.rows:
        if str(row[0].value) == str(class_id) or str(row[2].value) == str(class_id):
            found = True
            return row[4].value
    if not found:
        return None


def xuly():
    for i in range(2, tkb.max_row + 1):
        week = tkb.cell(row=i, column=5).value
        if isinstance(week, int):
            data['Subject'].append(get_name(tkb.cell(row=i, column=9).value))
            start_date = tuan1 + datetime.timedelta(weeks=week - 1, days=tkb.cell(row=i, column=1).value - 2)
            data['Start Date'].append(start_date.strftime("%d/%m/%Y"))
            data['End Date'].append(start_date.strftime("%d/%m/%Y"))
            times = tkb.cell(row=i, column=3).value.split('-')
            data['Start Time'].append(times[0])
            data['End Time'].append(times[1])
            data['Location'].append(tkb.cell(row=i, column=7).value)
            data['Description'].append(tkb.cell(row=i, column=9).value)
        else:
            weeks = week.split(',')
            for w in weeks:
                for k in range(int(w.split('-')[0]), int(w.split('-')[1]) + 1):
                    data['Subject'].append(get_name(tkb.cell(row=i, column=9).value))
                    start_date = tuan1 + datetime.timedelta(weeks=k - 1, days=int(tkb.cell(row=i, column=1).value) - 2)
                    data['Start Date'].append(start_date.strftime("%d/%m/%Y"))
                    data['End Date'].append(start_date.strftime("%d/%m/%Y"))
                    times = tkb.cell(row=i, column=3).value.split('-')
                    data['Start Time'].append(times[0])
                    data['End Time'].append(times[1])
                    data['Location'].append(tkb.cell(row=i, column=7).value)
                    data['Description'].append(tkb.cell(row=i, column=9).value)


xuly()
print(data)
result = pd.DataFrame(data)
result.to_csv('result.csv', index=False, encoding='utf-8-sig')
